import pandas as pd
import numpy as np
import openpyxl
import os
import shutil
import logging

from .config import DEPTH_COLUMN, LITHOLOGY_COLUMN, ANALYSIS_COLUMNS, RESEARCHED_LITHOLOGY_DEFAULTS, INVALID_DATA_VALUE, DEFAULT_MERGE_THRESHOLD

# Set up logging
logger = logging.getLogger(__name__)

class Analyzer:
    def __init__(self):
        pass

    def classify_rows(self, dataframe, lithology_rules, mnemonic_map, use_researched_defaults=True, use_fallback_classification=False):
        """
        Classifies rows in the DataFrame based on lithology rules, using specified mnemonics.

        Args:
            dataframe (pandas.DataFrame): The preprocessed DataFrame.
            lithology_rules (list): A list of dictionaries, each defining a lithology rule.
            mnemonic_map (dict): A dictionary mapping standardized names to original mnemonics
                                 (e.g., {'gamma': 'GR', 'density': 'RHOB', 'short_space_density': 'DENS', 'long_space_density': 'LSD'}).

        Returns:
            pandas.DataFrame: The DataFrame with a new 'LITHOLOGY_CODE' column.
        """
        classified_df = dataframe.copy()
        classified_df[LITHOLOGY_COLUMN] = 'NL'

        # Determine which columns to use for gamma and density based on mnemonic_map
        gamma_col_name = 'gamma' # Standardized name for gamma
        
        # Prioritize 'density', then 'short_space_density', then 'long_space_density'
        density_col_name = None
        if 'density' in mnemonic_map and mnemonic_map['density'] in classified_df.columns:
            density_col_name = 'density'
        elif 'short_space_density' in mnemonic_map and mnemonic_map['short_space_density'] in classified_df.columns:
            density_col_name = 'short_space_density'
        elif 'long_space_density' in mnemonic_map and mnemonic_map['long_space_density'] in classified_df.columns:
            density_col_name = 'long_space_density'
        
        if gamma_col_name not in classified_df.columns:
            logger.warning(f"Gamma column '{gamma_col_name}' not found in DataFrame. Classification may be inaccurate.")
            return classified_df # Cannot classify without gamma
        
        if density_col_name is None:
            logger.warning("No suitable density column found in DataFrame for classification. Classification may be inaccurate.")
            return classified_df # Cannot classify without density

        for rule in lithology_rules:
            name = rule.get('name')
            code = rule.get('code')
            
            # Get current rule values
            gamma_min = rule.get('gamma_min')
            gamma_max = rule.get('gamma_max')
            density_min = rule.get('density_min')
            density_max = rule.get('density_max')

            # Skip rule if it's for 'Not Logged' as it's a default classification
            if code == 'NL':
                continue

            # Determine if parameters are marked as "don't care" (both min and max are INVALID_DATA_VALUE)
            gamma_ignore = (gamma_min == INVALID_DATA_VALUE and gamma_max == INVALID_DATA_VALUE)
            density_ignore = (density_min == INVALID_DATA_VALUE and density_max == INVALID_DATA_VALUE)

            # Handle gamma and density ranges based on user preference for researched defaults
            if code in RESEARCHED_LITHOLOGY_DEFAULTS:
                researched_defaults = RESEARCHED_LITHOLOGY_DEFAULTS[code]

                if use_researched_defaults:
                    # Apply gamma defaults if current rule's gamma range is don't care OR zero
                    gamma_missing = gamma_ignore or (gamma_min == 0.0 and gamma_max == 0.0)
                    if gamma_missing and 'gamma_min' in researched_defaults and 'gamma_max' in researched_defaults:
                        gamma_min = researched_defaults['gamma_min']
                        gamma_max = researched_defaults['gamma_max']
                        gamma_ignore = False  # No longer ignore this parameter
                        logger.debug(f"Applying researched gamma defaults for {code}: {gamma_min}-{gamma_max}")

                    # Apply density defaults if current rule's density range is don't care OR zero
                    density_missing = density_ignore or (density_min == 0.0 and density_max == 0.0)
                    if density_missing and 'density_min' in researched_defaults and 'density_max' in researched_defaults:
                        density_min = researched_defaults['density_min']
                        density_max = researched_defaults['density_max']
                        density_ignore = False  # No longer ignore this parameter
                        logger.debug(f"Applying researched density defaults for {code}: {density_min}-{density_max}")

            # Create a boolean mask for rows that are not yet classified
            unclassified_mask = (classified_df[LITHOLOGY_COLUMN] == 'NL')

            # Create conditions for gamma and density, handling "don't care" markers
            gamma_condition = ((classified_df[gamma_col_name] >= gamma_min) & (classified_df[gamma_col_name] <= gamma_max)) if not gamma_ignore else True
            density_condition = ((classified_df[density_col_name] >= density_min) & (classified_df[density_col_name] <= density_max)) if not density_ignore else True

            # Create a boolean mask for the current rule
            rule_mask = gamma_condition & density_condition

            # Apply the rule only to unclassified rows that match the rule criteria
            classified_df.loc[unclassified_mask & rule_mask, LITHOLOGY_COLUMN] = code

        # Apply fallback classification for remaining 'NL' rows if enabled
        if use_fallback_classification:
            classified_df = self._classify_fallbacks(classified_df, gamma_col_name, density_col_name)

        return classified_df

    def _classify_fallbacks(self, dataframe, gamma_col_name, density_col_name):
        """
        Classify remaining 'NL' rows using fallback methods.

        Args:
            dataframe (pandas.DataFrame): DataFrame with 'NL' rows to classify
            gamma_col_name (str): Name of gamma column
            density_col_name (str): Name of density column

        Returns:
            pandas.DataFrame: DataFrame with fallback classifications applied
        """
        logger.debug("Applying fallback classification to 'NL' rows")

        # Find rows that are still 'NL'
        nl_mask = (dataframe[LITHOLOGY_COLUMN] == 'NL')
        nl_count = nl_mask.sum()

        if nl_count == 0:
            logger.debug("No 'NL' rows found - no fallback needed")
            return dataframe

        logger.debug(f"Found {nl_count} 'NL' rows for fallback classification")

        # Apply fallback using researched defaults
        fallback_classified_df = dataframe.copy()

        for idx in dataframe[nl_mask].index:
            row = dataframe.loc[idx]
            gamma_val = row[gamma_col_name]
            density_val = row[density_col_name]

            # Find best matching lithology from researched defaults
            best_match = self._get_nearest_lithology(gamma_val, density_val)

            if best_match:
                fallback_classified_df.loc[idx, LITHOLOGY_COLUMN] = best_match
                logger.debug(f"Fallback classified row {idx}: gamma={gamma_val:.1f}, density={density_val:.3f} -> {best_match}")

        # Apply extreme value rules for any remaining 'NL' rows
        remaining_nl_mask = (fallback_classified_df[LITHOLOGY_COLUMN] == 'NL')
        if remaining_nl_mask.sum() > 0:
            fallback_classified_df = self._apply_extreme_value_rules(fallback_classified_df, gamma_col_name, density_col_name, remaining_nl_mask)

        final_nl_count = (fallback_classified_df[LITHOLOGY_COLUMN] == 'NL').sum()
        logger.debug(f"Fallback classification complete. Remaining 'NL' rows: {final_nl_count}")

        return fallback_classified_df

    def _get_nearest_lithology(self, gamma_val, density_val):
        """
        Find the nearest lithology match using researched defaults.

        Args:
            gamma_val (float): Gamma ray value
            density_val (float): Density value

        Returns:
            str: Best matching lithology code, or None if no match found
        """
        best_match = None
        min_distance = float('inf')

        for code, defaults in RESEARCHED_LITHOLOGY_DEFAULTS.items():
            # Calculate distance in parameter space
            gamma_center = (defaults['gamma_min'] + defaults['gamma_max']) / 2
            density_center = (defaults['density_min'] + defaults['density_max']) / 2

            # Euclidean distance in normalized space
            gamma_range = defaults['gamma_max'] - defaults['gamma_min']
            density_range = defaults['density_max'] - defaults['density_min']

            if gamma_range > 0 and density_range > 0:
                gamma_distance = abs(gamma_val - gamma_center) / gamma_range
                density_distance = abs(density_val - density_center) / density_range
                distance = (gamma_distance ** 2 + density_distance ** 2) ** 0.5

                if distance < min_distance:
                    min_distance = distance
                    best_match = code

        # Only return match if it's reasonably close (within 2 standard deviations)
        if min_distance <= 2.0:
            return best_match

        return None

    def _apply_extreme_value_rules(self, dataframe, gamma_col_name, density_col_name, nl_mask):
        """
        Apply rules for extreme parameter values that don't match any standard lithologies.

        Args:
            dataframe (pandas.DataFrame): DataFrame to classify
            gamma_col_name (str): Name of gamma column
            density_col_name (str): Name of density column
            nl_mask (pandas.Series): Boolean mask for 'NL' rows

        Returns:
            pandas.DataFrame: DataFrame with extreme value classifications
        """
        classified_df = dataframe.copy()

        for idx in dataframe[nl_mask].index:
            gamma_val = dataframe.loc[idx, gamma_col_name]
            density_val = dataframe.loc[idx, density_col_name]

            # Extreme low density (gas, organic-rich)
            if density_val < 1.0:
                classified_df.loc[idx, LITHOLOGY_COLUMN] = 'CO'  # Coal
                logger.debug(f"Extreme low density fallback: row {idx}, density={density_val:.3f} -> CO")

            # Extreme high density (metamorphic, dense igneous)
            elif density_val > 3.5:
                classified_df.loc[idx, LITHOLOGY_COLUMN] = 'IG'  # Igneous (would need to add this rule)
                logger.debug(f"Extreme high density fallback: row {idx}, density={density_val:.3f} -> IG")

            # Extreme high gamma (very shaly, radioactive)
            elif gamma_val > 200:
                classified_df.loc[idx, LITHOLOGY_COLUMN] = 'SH'  # Shale
                logger.debug(f"Extreme high gamma fallback: row {idx}, gamma={gamma_val:.1f} -> SH")

            # Very low gamma, moderate density (clean sandstones or carbonates)
            elif gamma_val < 10 and 2.0 <= density_val <= 3.0:
                if density_val < 2.7:
                    classified_df.loc[idx, LITHOLOGY_COLUMN] = 'SS'  # Sandstone
                    logger.debug(f"Clean sandstone fallback: row {idx}, gamma={gamma_val:.1f}, density={density_val:.3f} -> SS")
                else:
                    classified_df.loc[idx, LITHOLOGY_COLUMN] = 'LS'  # Limestone (would need to add this rule)
                    logger.debug(f"Carbonate fallback: row {idx}, gamma={gamma_val:.1f}, density={density_val:.3f} -> LS")

        return classified_df

    def classify_rows_simple(self, dataframe, lithology_rules, mnemonic_map):
        """
        Classifies rows using the Simple method: first by density only, then by gamma only.
        
        Args:
            dataframe (pandas.DataFrame): The preprocessed DataFrame.
            lithology_rules (list): A list of dictionaries, each defining a lithology rule.
            mnemonic_map (dict): A dictionary mapping standardized names to original mnemonics.

        Returns:
            pandas.DataFrame: The DataFrame with a new 'LITHOLOGY_CODE' column.
        """
        classified_df = dataframe.copy()
        classified_df[LITHOLOGY_COLUMN] = 'NL'

        # Determine which columns to use for gamma and density based on mnemonic_map
        gamma_col_name = 'gamma' # Standardized name for gamma
        
        # Prioritize 'density', then 'short_space_density', then 'long_space_density'
        density_col_name = None
        if 'density' in mnemonic_map and mnemonic_map['density'] in classified_df.columns:
            density_col_name = 'density'
        elif 'short_space_density' in mnemonic_map and mnemonic_map['short_space_density'] in classified_df.columns:
            density_col_name = 'short_space_density'
        elif 'long_space_density' in mnemonic_map and mnemonic_map['long_space_density'] in classified_df.columns:
            density_col_name = 'long_space_density'
        
        if gamma_col_name not in classified_df.columns:
            logger.warning(f"Gamma column '{gamma_col_name}' not found in DataFrame. Classification may be inaccurate.")
            return classified_df # Cannot classify without gamma
        
        if density_col_name is None:
            logger.warning("No suitable density column found in DataFrame for classification. Classification may be inaccurate.")
            return classified_df # Cannot classify without density

        # First pass: classify by density only (ignore gamma)
        for rule in lithology_rules:
            name = rule.get('name')
            code = rule.get('code')
            
            # Skip rule if it's for 'Not Logged' as it's a default classification
            if code == 'NL':
                continue

            # Get current rule values
            density_min = rule.get('density_min')
            density_max = rule.get('density_max')

            # Skip rules with invalid density ranges
            if density_min == INVALID_DATA_VALUE and density_max == INVALID_DATA_VALUE:
                continue

            # Create a boolean mask for rows that are not yet classified
            unclassified_mask = (classified_df[LITHOLOGY_COLUMN] == 'NL')

            # Create condition for density only
            density_condition = ((classified_df[density_col_name] >= density_min) & 
                               (classified_df[density_col_name] <= density_max))

            # Apply the rule only to unclassified rows that match the density criteria
            classified_df.loc[unclassified_mask & density_condition, LITHOLOGY_COLUMN] = code

        # Second pass: classify by gamma only (overwrite previous classifications)
        for rule in lithology_rules:
            name = rule.get('name')
            code = rule.get('code')
            
            # Skip rule if it's for 'Not Logged' as it's a default classification
            if code == 'NL':
                continue

            # Get current rule values
            gamma_min = rule.get('gamma_min')
            gamma_max = rule.get('gamma_max')

            # Skip rules with invalid gamma ranges (ignore -999.25 values)
            if gamma_min == INVALID_DATA_VALUE and gamma_max == INVALID_DATA_VALUE:
                continue

            # Create condition for gamma only
            gamma_condition = ((classified_df[gamma_col_name] >= gamma_min) & 
                             (classified_df[gamma_col_name] <= gamma_max))

            # Apply the rule to ALL rows that match the gamma criteria (overwrites previous classifications)
            classified_df.loc[gamma_condition, LITHOLOGY_COLUMN] = code

        return classified_df

    def group_into_units(self, dataframe, lithology_rules, smart_interbedding=False, smart_interbedding_max_sequence_length=10, smart_interbedding_thick_unit_threshold=0.5):
        """
        Groups contiguous blocks of rows with the same LITHOLOGY_CODE into lithological units.
        If smart_interbedding is enabled, detects alternating lithologies and groups them into interbedded units.

        Args:
            dataframe (pandas.DataFrame): The classified DataFrame.
            lithology_rules (list): A list of dictionaries, each defining a lithology rule.
            smart_interbedding (bool): Whether to enable smart interbedding detection.
            smart_interbedding_max_sequence_length (int): Maximum segments in interbedded sequence before stopping.
            smart_interbedding_thick_unit_threshold (float): Stop interbedding when next unit exceeds this thickness (meters).

        Returns:
            pandas.DataFrame: A DataFrame summarizing the lithological units.
        """
        logger.debug(f"group_into_units: DataFrame columns: {dataframe.columns.tolist()}")
        logger.debug(f"group_into_units: DataFrame index: {dataframe.index.tolist()}")

        if LITHOLOGY_COLUMN not in dataframe.columns:
            raise ValueError(f"DataFrame must contain a '{LITHOLOGY_COLUMN}' column.")
        if DEPTH_COLUMN not in dataframe.columns:
            raise ValueError(f"Depth column '{DEPTH_COLUMN}' not found in DataFrame.")

        # Create a mapping from lithology code to rule details
        rules_map = {rule['code']: rule for rule in lithology_rules}

        # Ensure the DataFrame is sorted by depth for correct grouping
        sorted_df = dataframe.sort_values(by=DEPTH_COLUMN).reset_index(drop=True)

        # Always use standard grouping - smart interbedding now runs as post-processing
        return self._group_standard_units(sorted_df, rules_map)

    def find_interbedding_candidates(self, units_df, max_sequence_length=10, thick_unit_threshold=0.5):
        """
        Scan units dataframe for potential interbedding candidates.
        This runs as post-processing after analysis is complete.

        Args:
            units_df (pandas.DataFrame): DataFrame of lithological units
            max_sequence_length (int): Maximum number of alternating units to consider
            thick_unit_threshold (float): Skip interbedding detection for units thicker than this

        Returns:
            list: List of interbedding candidate dictionaries
        """
        print(f"DEBUG: find_interbedding_candidates called with max_sequence_length={max_sequence_length}, thick_unit_threshold={thick_unit_threshold}")
        print(f"DEBUG: units_df shape: {units_df.shape if hasattr(units_df, 'shape') else 'No shape'}")
        print(f"DEBUG: units_df columns: {list(units_df.columns) if hasattr(units_df, 'columns') else 'No columns'}")

        if units_df.empty or len(units_df) <= 1:
            print("DEBUG: units_df is empty or has <= 1 rows, returning empty candidates")
            return []

        candidates = []
        i = 0
        total_iterations = 0

        print(f"DEBUG: Starting scan through {len(units_df)} units")
        while i < len(units_df) and total_iterations < 1000:  # Safety limit
            print(f"DEBUG: Iteration {total_iterations}, checking from index {i}")
            # Look for alternating pattern starting from current position
            candidate = self._find_interbedding_candidate(units_df, i, max_sequence_length, thick_unit_threshold)

            if candidate:
                print(f"DEBUG: Found candidate at index {i}: {candidate.get('from_depth')} - {candidate.get('to_depth')}, {len(candidate.get('lithologies', []))} lithologies")
                candidates.append(candidate)
                # Skip the units that were included in this candidate
                units_skipped = len(candidate['original_sequence'])
                print(f"DEBUG: Skipping {units_skipped} units after finding candidate")
                i += units_skipped
            else:
                print(f"DEBUG: No candidate found at index {i}, moving to next unit")
                i += 1

            total_iterations += 1

        print(f"DEBUG: Scan complete. Found {len(candidates)} interbedding candidates")
        for idx, cand in enumerate(candidates):
            print(f"DEBUG: Candidate {idx}: from_depth={cand.get('from_depth')}, to_depth={cand.get('to_depth')}, lithologies={[l.get('code') for l in cand.get('lithologies', [])]}")

        logger.debug(f"Found {len(candidates)} interbedding candidates")
        return candidates

    def _find_interbedding_candidate(self, units_df, start_idx, max_sequence_length=10, thick_unit_threshold=0.5):
        """
        Find a single interbedding candidate starting from the given index.

        Args:
            units_df (pandas.DataFrame): DataFrame of lithological units
            start_idx (int): Starting index in the units dataframe
            max_sequence_length (int): Maximum number of alternating units to consider
            thick_unit_threshold (float): Skip interbedding detection for units thicker than this

        Returns:
            dict: Interbedding candidate dictionary or None if no candidate found
        """
        print(f"DEBUG: _find_interbedding_candidate called with start_idx={start_idx}, max_sequence_length={max_sequence_length}, thick_unit_threshold={thick_unit_threshold}")

        if start_idx >= len(units_df):
            print("DEBUG: start_idx >= len(units_df), returning None")
            return None

        # Get sequence of alternating units
        print(f"DEBUG: Calling _extract_alternating_sequence from index {start_idx}")
        sequence = self._extract_alternating_sequence(units_df, start_idx, max_sequence_length, thick_unit_threshold)
        print(f"DEBUG: _extract_alternating_sequence returned {len(sequence) if sequence else 0} units")

        if not sequence or len(sequence) < 3:  # Need at least 3 units for meaningful interbedding
            print(f"DEBUG: Sequence too short (length={len(sequence) if sequence else 0}), need at least 3 units. Returning None")
            return None

        # Calculate metrics for the sequence
        total_thickness = sum(unit['thickness'] for unit in sequence)
        print(f"DEBUG: Total thickness of sequence: {total_thickness}")

        # Calculate average layer thickness (total thickness ÷ number of layers)
        # This matches the user's specification: "the layer thickness calculation should be a sum of all grouped lithology units that are creating the interbedded section"
        avg_layer_thickness = total_thickness / len(sequence)
        print(f"DEBUG: Average layer thickness: {avg_layer_thickness}")

        # Only consider for interbedding if average layer thickness < 200mm
        if avg_layer_thickness >= 0.2:
            print(f"DEBUG: Average layer thickness {avg_layer_thickness} >= 0.2m, not considering for interbedding")
            return None

        print("DEBUG: Average layer thickness < 0.2m, proceeding with interbedding analysis")

        # Determine interrelationship code based on average layer thickness
        if avg_layer_thickness < 0.02:
            inter_code = 'IL'  # Interlaminated (< 20mm)
        elif avg_layer_thickness < 0.06:
            inter_code = 'UB'  # Very Thinly Interbedded (20-60mm)
        elif avg_layer_thickness < 0.2:
            inter_code = 'TB'  # Thinly Interbedded (60-200mm)
        else:
            inter_code = 'CB'  # Coarsely Interbedded (> 200mm)

        print(f"DEBUG: Determined interrelationship code: {inter_code}")

        # Calculate lithology percentages and dominance
        lithology_thicknesses = {}
        for unit in sequence:
            code = unit[LITHOLOGY_COLUMN]
            lithology_thicknesses[code] = lithology_thicknesses.get(code, 0) + unit['thickness']

        print(f"DEBUG: Lithology thicknesses: {lithology_thicknesses}")

        # Sort by thickness (dominance) - user specified "by total thickness"
        sorted_lithologies = sorted(lithology_thicknesses.items(), key=lambda x: x[1], reverse=True)
        print(f"DEBUG: Sorted lithologies by thickness: {sorted_lithologies}")

        # Create lithology components with percentages and sequence numbers
        lithologies = []
        for seq_num, (code, thickness) in enumerate(sorted_lithologies, 1):
            percentage = (thickness / total_thickness) * 100
            print(f"DEBUG: Lithology {code}: thickness={thickness}, percentage={percentage}")

            # Skip lithologies < 5% unless they are the dominant one
            if seq_num > 1 and percentage < 5:
                print(f"DEBUG: Skipping lithology {code} (percentage {percentage} < 5% and not dominant)")
                continue

            lithologies.append({
                'code': code,
                'thickness': thickness,
                'percentage': round(percentage, 2),
                'sequence': seq_num
            })

        print(f"DEBUG: Final lithologies list: {[l['code'] for l in lithologies]}")

        # Only proceed if we have at least 2 lithologies after filtering
        if len(lithologies) < 2:
            print(f"DEBUG: Only {len(lithologies)} lithologies after filtering, need at least 2. Returning None")
            return None

        # Create candidate dictionary
        candidate = {
            'from_depth': sequence[0]['from_depth'],
            'to_depth': sequence[-1]['to_depth'],
            'original_sequence': sequence,
            'lithologies': lithologies,
            'average_layer_thickness': avg_layer_thickness,
            'interrelationship_code': inter_code,
            'total_thickness': total_thickness
        }

        print(f"DEBUG: Created candidate: from_depth={candidate['from_depth']}, to_depth={candidate['to_depth']}")
        return candidate

    def _extract_alternating_sequence(self, units_df, start_idx, max_sequence_length=10, thick_unit_threshold=0.5):
        """
        Extract a sequence of alternating lithology units.

        Args:
            units_df (pandas.DataFrame): DataFrame of lithological units
            start_idx (int): Starting index
            max_sequence_length (int): Maximum sequence length
            thick_unit_threshold (float): Stop if unit exceeds this thickness

        Returns:
            list: List of unit dictionaries in the alternating sequence
        """
        print(f"DEBUG: _extract_alternating_sequence called with start_idx={start_idx}, max_sequence_length={max_sequence_length}, thick_unit_threshold={thick_unit_threshold}")

        if start_idx >= len(units_df):
            print("DEBUG: start_idx >= len(units_df), returning empty sequence")
            return []

        sequence = []
        current_code = None
        units_added = 0

        print(f"DEBUG: Scanning from index {start_idx} to {min(start_idx + max_sequence_length, len(units_df))}")

        for i in range(start_idx, min(start_idx + max_sequence_length, len(units_df))):
            unit = units_df.iloc[i]
            unit_code = unit[LITHOLOGY_COLUMN]
            unit_thickness = unit['thickness']

            print(f"DEBUG: Checking unit at index {i}: code={unit_code}, thickness={unit_thickness}")

            # Skip units that are too thick (user's thick unit threshold)
            if unit_thickness > thick_unit_threshold:
                print(f"DEBUG: Unit thickness {unit_thickness} > thick_unit_threshold {thick_unit_threshold}, stopping sequence extraction")
                break

            # If this is a different lithology than the previous one, add it
            if unit_code != current_code:
                print(f"DEBUG: Adding unit with code {unit_code} (different from previous {current_code})")
                sequence.append(unit.to_dict())
                current_code = unit_code
                units_added += 1

                # Stop if we've added too many units
                if units_added >= max_sequence_length:
                    print(f"DEBUG: Reached max_sequence_length {max_sequence_length}, stopping")
                    break
            else:
                print(f"DEBUG: Same lithology {unit_code} as previous, breaking alternating pattern")
                # Same lithology - this breaks the alternating pattern
                break

        print(f"DEBUG: Extracted sequence with {len(sequence)} units: {[u[LITHOLOGY_COLUMN] for u in sequence]}")
        return sequence

    def apply_interbedding_candidates(self, units_df, candidates, selected_indices, lithology_rules):
        """
        Apply selected interbedding candidates to create interbedded units.

        Args:
            units_df (pandas.DataFrame): Original units dataframe
            candidates (list): List of all candidate dictionaries
            selected_indices (list): List of indices of selected candidates
            lithology_rules (list): List of lithology rules for rule lookup

        Returns:
            pandas.DataFrame: Updated units dataframe with interbedded units
        """
        if not selected_indices:
            return units_df

        # Create rules map for lookup
        rules_map = {rule['code']: rule for rule in lithology_rules}

        # Sort selected candidates by starting depth (process in order)
        candidates_to_apply = [candidates[idx] for idx in sorted(selected_indices)]

        updated_units = []
        units_processed = 0

        for candidate in candidates_to_apply:
            # Find the range of units this candidate covers
            candidate_start_depth = candidate['from_depth']
            candidate_end_depth = candidate['to_depth']

            # Add any units before this candidate
            while units_processed < len(units_df):
                unit = units_df.iloc[units_processed]
                if unit['from_depth'] >= candidate_start_depth:
                    break
                updated_units.append(unit.to_dict())
                units_processed += 1

            # Create interbedded units for this candidate
            for lithology in candidate['lithologies']:
                rule = rules_map.get(lithology['code'], {})

                interbedded_unit = {
                    'from_depth': candidate_start_depth,
                    'to_depth': candidate_end_depth,
                    'thickness': candidate_end_depth - candidate_start_depth,  # Full section thickness
                    LITHOLOGY_COLUMN: lithology['code'],
                    'lithology_qualifier': rule.get('qualifier', ''),
                    'shade': rule.get('shade', ''),
                    'hue': rule.get('hue', ''),
                    'colour': rule.get('colour', ''),
                    'weathering': rule.get('weathering', ''),
                    'estimated_strength': rule.get('strength', ''),
                    'background_color': rule.get('background_color', '#FFFFFF'),
                    'svg_path': rule.get('svg_path'),
                    'record_sequence': lithology['sequence'],
                    'inter_relationship': candidate['interrelationship_code'] if lithology['sequence'] == 1 else '',
                    'percentage': lithology['percentage']
                }
                updated_units.append(interbedded_unit)

            # Skip the original units that were replaced by interbedding
            while units_processed < len(units_df):
                unit = units_df.iloc[units_processed]
                if unit['to_depth'] > candidate_end_depth:
                    break
                units_processed += 1

        # Add any remaining units
        while units_processed < len(units_df):
            updated_units.append(units_df.iloc[units_processed].to_dict())
            units_processed += 1

        # Convert back to DataFrame
        result_df = pd.DataFrame(updated_units)

        # Ensure proper column ordering
        if not result_df.empty:
            result_df = result_df[[
                'from_depth', 'to_depth', 'thickness', LITHOLOGY_COLUMN,
                'lithology_qualifier', 'shade', 'hue', 'colour',
                'weathering', 'estimated_strength', 'background_color', 'svg_path',
                'record_sequence', 'inter_relationship', 'percentage'
            ]]

        return result_df

    def _group_standard_units(self, sorted_df, rules_map):
        """Standard unit grouping without interbedding detection."""
        units = []
        current_unit = None

        for i, row in sorted_df.iterrows():
            lithology_code = row[LITHOLOGY_COLUMN]
            current_depth = row[DEPTH_COLUMN]
            rule = rules_map.get(lithology_code, {})

            if current_unit is None:
                # Start a new unit
                current_unit = {
                    'from_depth': current_depth,
                    'to_depth': current_depth,
                    LITHOLOGY_COLUMN: lithology_code,
                    'lithology_qualifier': rule.get('qualifier', ''),
                    'shade': rule.get('shade', ''),
                    'hue': rule.get('hue', ''),
                    'colour': rule.get('colour', ''),
                    'weathering': rule.get('weathering', ''),
                    'estimated_strength': rule.get('strength', ''),
                    'background_color': rule.get('background_color', '#FFFFFF'),
                    'svg_path': rule.get('svg_path'),
                    'record_sequence': '',  # New column for interbedding
                    'inter_relationship': '',  # New column for interbedding
                    'percentage': 0.0  # New column for interbedding
                }
            elif lithology_code == current_unit[LITHOLOGY_COLUMN]:
                # Continue the current unit
                current_unit['to_depth'] = current_depth
            else:
                # End the previous unit and start a new one
                current_unit['to_depth'] = current_depth
                units.append(current_unit)

                # Start a new unit
                current_unit = {
                    'from_depth': current_depth,
                    'to_depth': current_depth,
                    LITHOLOGY_COLUMN: lithology_code,
                    'lithology_qualifier': rule.get('qualifier', ''),
                    'shade': rule.get('shade', ''),
                    'hue': rule.get('hue', ''),
                    'colour': rule.get('colour', ''),
                    'weathering': rule.get('weathering', ''),
                    'estimated_strength': rule.get('strength', ''),
                    'background_color': rule.get('background_color', '#FFFFFF'),
                    'svg_path': rule.get('svg_path'),
                    'record_sequence': '',  # New column for interbedding
                    'inter_relationship': '',  # New column for interbedding
                    'percentage': 0.0  # New column for interbedding
                }

        # Add the last unit
        if current_unit is not None:
            current_unit['to_depth'] = sorted_df.iloc[-1][DEPTH_COLUMN]
            units.append(current_unit)

        # Convert to DataFrame
        units_df = pd.DataFrame(units)

        # Calculate thickness and reorder columns
        if not units_df.empty:
            units_df.loc[:, 'thickness'] = units_df['to_depth'] - units_df['from_depth']
            units_df = units_df[[
                'from_depth', 'to_depth', 'thickness', LITHOLOGY_COLUMN,
                'lithology_qualifier', 'shade', 'hue', 'colour',
                'weathering', 'estimated_strength', 'background_color', 'svg_path',
                'record_sequence', 'inter_relationship', 'percentage'
            ]]
        else:
            units_df = pd.DataFrame(columns=[
                'from_depth', 'to_depth', 'thickness', LITHOLOGY_COLUMN,
                'lithology_qualifier', 'shade', 'hue', 'colour',
                'weathering', 'estimated_strength', 'background_color', 'svg_path',
                'record_sequence', 'inter_relationship', 'percentage'
            ])

        return units_df

    def _group_with_smart_interbedding(self, sorted_df, rules_map, max_sequence_length=10, thick_unit_threshold=0.5):
        """Group units with smart interbedding detection."""
        units = []

        logger.debug(f"Starting smart interbedding detection on {len(sorted_df)} rows")

        # First pass: identify potential interbedded sequences
        i = 0
        interbedding_count = 0
        while i < len(sorted_df):
            current_code = sorted_df.iloc[i][LITHOLOGY_COLUMN]

            # Look ahead to find alternating pattern
            interbedded_sequence = self._find_interbedded_sequence(sorted_df, i, max_sequence_length, thick_unit_threshold)

            if interbedded_sequence:
                logger.debug(f"Found interbedded sequence at index {i}: {len(interbedded_sequence)} segments")
                # Process interbedded sequence
                interbedded_units = self._process_interbedded_sequence(sorted_df, interbedded_sequence, rules_map)
                units.extend(interbedded_units)
                interbedding_count += 1
                i = interbedded_sequence[-1]['end_index'] + 1
            else:
                # Process as regular unit
                regular_unit = self._create_regular_unit(sorted_df, i, rules_map)
                units.append(regular_unit)
                i += 1

        logger.debug(f"Smart interbedding completed: {interbedding_count} interbedded sequences found, {len(units)} total units created")

        # Convert to DataFrame
        units_df = pd.DataFrame(units)

        # Calculate thickness and reorder columns
        if not units_df.empty:
            units_df.loc[:, 'thickness'] = units_df['to_depth'] - units_df['from_depth']
            units_df = units_df[[
                'from_depth', 'to_depth', 'thickness', LITHOLOGY_COLUMN,
                'lithology_qualifier', 'shade', 'hue', 'colour',
                'weathering', 'estimated_strength', 'background_color', 'svg_path',
                'record_sequence', 'inter_relationship', 'percentage'
            ]]
        else:
            units_df = pd.DataFrame(columns=[
                'from_depth', 'to_depth', 'thickness', LITHOLOGY_COLUMN,
                'lithology_qualifier', 'shade', 'hue', 'colour',
                'weathering', 'estimated_strength', 'background_color', 'svg_path',
                'record_sequence', 'inter_relationship', 'percentage'
            ])

        return units_df

    def _find_interbedded_sequence(self, sorted_df, start_idx, max_sequence_length=10, thick_unit_threshold=0.5):
        """Find a sequence of alternating lithologies that should be interbedded."""
        if start_idx >= len(sorted_df):
            return None

        # Get initial lithology
        initial_code = sorted_df.iloc[start_idx][LITHOLOGY_COLUMN]
        sequence = []
        current_code = initial_code
        idx = start_idx

        # Look for alternating pattern
        while idx < len(sorted_df):
            row = sorted_df.iloc[idx]
            code = row[LITHOLOGY_COLUMN]

            # Count consecutive rows with same lithology
            count = 0
            start_depth = row[DEPTH_COLUMN]

            while idx < len(sorted_df) and sorted_df.iloc[idx][LITHOLOGY_COLUMN] == code:
                count += 1
                idx += 1

            end_depth = sorted_df.iloc[idx-1][DEPTH_COLUMN] if idx > 0 else start_depth

            sequence.append({
                'code': code,
                'count': count,
                'start_index': idx - count,
                'end_index': idx - 1,
                'start_depth': start_depth,
                'end_depth': end_depth,
                'thickness': end_depth - start_depth
            })

            # Check if we have enough alternation cycles and meet criteria
            if len(sequence) >= 4:  # At least 2 full alternation cycles
                avg_thickness = sum(s['thickness'] for s in sequence) / len(sequence)

                # Check if average thickness < 200mm and we have alternating pattern
                if avg_thickness < 0.2 and self._is_alternating_sequence(sequence):
                    # Check for stop conditions
                    if idx < len(sorted_df):
                        next_code = sorted_df.iloc[idx][LITHOLOGY_COLUMN]
                        next_thickness = 0

                        # Check if next unit exceeds thick unit threshold (stop condition)
                        next_count = 0
                        while idx + next_count < len(sorted_df) and sorted_df.iloc[idx + next_count][LITHOLOGY_COLUMN] == next_code:
                            next_count += 1

                        if next_count > 0:
                            next_end_depth = sorted_df.iloc[idx + next_count - 1][DEPTH_COLUMN]
                            next_start_depth = sorted_df.iloc[idx][DEPTH_COLUMN]
                            next_thickness = next_end_depth - next_start_depth

                        if next_thickness > thick_unit_threshold:  # Configurable thick unit stop condition
                            break

                    return sequence

            # Stop if sequence gets too long or we hit a thick unit
            if len(sequence) > max_sequence_length or (sequence and sequence[-1]['thickness'] > thick_unit_threshold):
                break

        return None

    def _is_alternating_sequence(self, sequence):
        """Check if sequence represents alternating lithologies."""
        if len(sequence) < 4:  # Need at least 2 full cycles
            return False

        # Get unique codes in sequence
        codes = [s['code'] for s in sequence]
        unique_codes = list(set(codes))

        if len(unique_codes) < 2:
            return False  # Not alternating if only one lithology

        # For simple alternating between 2 lithologies
        if len(unique_codes) == 2:
            # Check if it alternates between the two
            expected_pattern = [unique_codes[0], unique_codes[1]] * (len(sequence) // 2)
            if len(sequence) % 2 == 1:
                expected_pattern.append(unique_codes[0])
            return codes == expected_pattern

        # For more than 2 lithologies, ensure no lithology repeats consecutively
        # This prevents patterns like A-A-B-C which are not truly alternating
        for i in range(1, len(codes)):
            if codes[i] == codes[i-1]:
                return False  # Consecutive same lithology found

        # Ensure we have alternation (no more than 2 consecutive different lithologies in a row if more than 2 types)
        # This is a basic check - more sophisticated logic could be added later
        return True

    def _process_interbedded_sequence(self, sorted_df, sequence, rules_map):
        """Process an interbedded sequence into multiple unit rows."""
        units = []

        # Calculate total thickness of interbedded section
        total_thickness = sum(s['thickness'] for s in sequence)

        # Calculate average layer thickness for interrelationship code
        avg_thickness = total_thickness / len(sequence)

        # Determine interrelationship code
        if avg_thickness < 0.02:
            inter_code = 'IL'  # Interlaminated
        elif avg_thickness < 0.06:
            inter_code = 'UB'  # Very Thinly Interbedded
        elif avg_thickness < 0.2:
            inter_code = 'TB'  # Thinly Interbedded
        else:
            inter_code = 'CB'  # Coarsely Interbedded

        # Get lithology counts for dominance
        code_counts = {}
        for s in sequence:
            code_counts[s['code']] = code_counts.get(s['code'], 0) + s['thickness']

        # Sort by thickness (dominance)
        sorted_codes = sorted(code_counts.items(), key=lambda x: x[1], reverse=True)

        # Create units for each lithology component
        from_depth = sequence[0]['start_depth']
        to_depth = sequence[-1]['end_depth']

        for seq_num, (code, thickness) in enumerate(sorted_codes, 1):
            percentage = (thickness / total_thickness) * 100

            # Skip if less than 5% and not the dominant lithology
            if seq_num > 1 and percentage < 5:
                continue

            rule = rules_map.get(code, {})

            unit = {
                'from_depth': from_depth,
                'to_depth': to_depth,
                LITHOLOGY_COLUMN: code,
                'lithology_qualifier': rule.get('qualifier', ''),
                'shade': rule.get('shade', ''),
                'hue': rule.get('hue', ''),
                'colour': rule.get('colour', ''),
                'weathering': rule.get('weathering', ''),
                'estimated_strength': rule.get('strength', ''),
                'background_color': rule.get('background_color', '#FFFFFF'),
                'svg_path': rule.get('svg_path'),
                'record_sequence': seq_num,
                'inter_relationship': inter_code if seq_num == 1 else '',
                'percentage': round(percentage, 2)
            }
            units.append(unit)

        return units

    def _create_regular_unit(self, sorted_df, idx, rules_map):
        """Create a regular (non-interbedded) unit."""
        row = sorted_df.iloc[idx]
        lithology_code = row[LITHOLOGY_COLUMN]
        rule = rules_map.get(lithology_code, {})

        return {
            'from_depth': row[DEPTH_COLUMN],
            'to_depth': row[DEPTH_COLUMN],
            LITHOLOGY_COLUMN: lithology_code,
            'lithology_qualifier': rule.get('qualifier', ''),
            'shade': rule.get('shade', ''),
            'hue': rule.get('hue', ''),
            'colour': rule.get('colour', ''),
            'weathering': rule.get('weathering', ''),
            'estimated_strength': rule.get('strength', ''),
            'background_color': rule.get('background_color', '#FFFFFF'),
            'svg_path': rule.get('svg_path'),
            'record_sequence': '',
            'inter_relationship': '',
            'percentage': 0.0
        }

    def merge_thin_units(self, units_df, threshold=DEFAULT_MERGE_THRESHOLD):
        """
        Merges lithological units that are thinner than the specified threshold and have the same lithology.

        Args:
            units_df (pandas.DataFrame): DataFrame of lithological units from group_into_units.
            threshold (float): Thickness threshold below which units are considered "thin" (default 0.05 meters = 5cm).

        Returns:
            pandas.DataFrame: DataFrame with thin adjacent units of same lithology merged.
        """
        if units_df.empty or len(units_df) <= 1:
            return units_df

        # Ensure units are sorted by depth
        merged_units = units_df.sort_values('from_depth').reset_index(drop=True)

        # List to hold final merged units
        final_units = []
        i = 0

        while i < len(merged_units):
            current_unit = merged_units.iloc[i].copy()

            # Check if this unit is thin and if there are more units to potentially merge
            while (current_unit['thickness'] < threshold and
                   i + 1 < len(merged_units)):

                next_unit = merged_units.iloc[i + 1]

                # Check if next unit has same lithology (code and qualifier)
                same_lithology = (current_unit[LITHOLOGY_COLUMN] == next_unit[LITHOLOGY_COLUMN] and
                                current_unit.get('lithology_qualifier', '') == next_unit.get('lithology_qualifier', ''))

                if same_lithology:
                    # Merge: extend current unit to include next unit
                    current_unit['to_depth'] = next_unit['to_depth']
                    current_unit['thickness'] = current_unit['to_depth'] - current_unit['from_depth']
                    i += 1  # Skip the merged unit
                else:
                    # Cannot merge, break the inner loop
                    break

            # Add the (potentially merged) unit to final list
            final_units.append(current_unit)
            i += 1

        # Convert back to DataFrame
        result_df = pd.DataFrame(final_units)

        # Ensure proper column ordering
        if not result_df.empty:
            result_df = result_df[[
                'from_depth', 'to_depth', 'thickness', LITHOLOGY_COLUMN,
                'lithology_qualifier', 'shade', 'hue', 'colour',
                'weathering', 'estimated_strength', 'background_color', 'svg_path',
                'record_sequence', 'inter_relationship', 'percentage'
            ]]

        return result_df

    def save_to_template(self, classified_data, template_path, output_path, callback=None, units=None):
        """
        Save classified lithology data to the 'Lithology' sheet in the TEMPLATE.xlsx file.
        
        Args:
            classified_data (pd.DataFrame): The dataframe with classified lithology
            template_path (str): Path to the template Excel file
            output_path (str): Path where the output file should be saved
            callback (function, optional): Callback function for progress updates
            units (pd.DataFrame, optional): Lithology units with from/to depths and properties
            
        Returns:
            bool: True if successful, False otherwise
        """
        try:
            # Update progress
            if callback:
                callback(f"Loading template file: {template_path}")
                
            # Check if the template file exists
            if not os.path.exists(template_path):
                error_msg = f"Template file not found: {template_path}"
                logger.error(error_msg)
                if callback:
                    callback(error_msg)
                return False
            
            # Check if template and output paths are the same
            if os.path.abspath(template_path) != os.path.abspath(output_path):
                # If not the same, make a copy of the template file to avoid modifying the original
                if callback:
                    callback(f"Creating a copy of the template file at: {output_path}")
                
                # Ensure output directory exists
                os.makedirs(os.path.dirname(output_path), exist_ok=True)
                
                # Copy the template to the output path
                shutil.copy2(template_path, output_path)
                
                # Load the copied workbook
                workbook = openpyxl.load_workbook(output_path)
            else:
                # If template and output paths are the same, load the workbook directly
                if callback:
                    callback(f"Writing directly to template file: {output_path}")
                workbook = openpyxl.load_workbook(output_path)
            
            # Try to get the 'Lithology' sheet
            if 'Lithology' in workbook.sheetnames:
                sheet = workbook['Lithology']
                if callback:
                    callback("Using 'Lithology' sheet in template")
            else:
                # If Lithology sheet doesn't exist, create it
                sheet = workbook.create_sheet('Lithology')
                if callback:
                    callback("Created new 'Lithology' sheet in template")
            
            # Update progress
            if callback:
                callback("Preparing data for template...")
            
            # Define starting row for data insertion (row 5 as specified)
            start_row = 5
            
            # Column mapping as specified
            column_mapping = {
                'from_depth': 'A',
                'to_depth': 'B',
                'thickness': 'D',
                LITHOLOGY_COLUMN: 'L',
                'lithology_qualifier': 'M', # Add lithology_qualifier to column mapping
                'shade': 'N',
                'hue': 'O',
                'colour': 'P',
                'weathering': 'Q',
                'estimated_strength': 'R',
                'record_sequence': 'S',  # New column for interbedding
                'inter_relationship': 'T',  # New column for interbedding
                'percentage': 'U'  # New column for interbedding
            }
            
            # Function to safely write to a cell, handling merged cells
            def safe_write_cell(sheet, cell_coord, value):
                # Check if the cell coordinate is valid
                try:
                    cell = sheet[cell_coord]
                    
                    # Check if it's a merged cell
                    for merged_range in sheet.merged_cells.ranges:
                        if cell.coordinate in merged_range:
                            # Skip this cell - don't try to write to merged cells
                            return False
                    
                    # If not a merged cell, write the value
                    cell.value = value
                    return True
                except Exception as e:
                    logger.warning(f"Couldn't write to cell {cell_coord}: {str(e)}")
                    return False
            
            # If we have units data, use it for the template
            if units is not None and not units.empty:
                if callback:
                    callback(f"Writing {len(units)} lithology units to Lithology sheet starting at row {start_row}...")

                # Debug: Log first 5 units being written
                if callback and len(units) > 0:
                     callback("DEBUG (save_to_template): First 5 units to be written:")
                     for idx, unit_debug in units.head(5).iterrows():
                         callback(f"  Index {idx}: {unit_debug.to_dict()}")

                # Write units to the template
                for i, unit in units.iterrows():
                    row_num = start_row + i

                    # Write from depth in column A
                    safe_write_cell(sheet, f'{column_mapping["from_depth"]}{row_num}', unit['from_depth'])
                    
                    # Write to depth in column B
                    safe_write_cell(sheet, f'{column_mapping["to_depth"]}{row_num}', unit['to_depth'])
                    
                    # Write thickness in column D
                    safe_write_cell(sheet, f'{column_mapping["thickness"]}{row_num}', unit['thickness'])
                    
                    # Write lithology code in column L
                    safe_write_cell(sheet, f'{column_mapping[LITHOLOGY_COLUMN]}{row_num}', unit[LITHOLOGY_COLUMN])
                    
                    # Write lithology qualifier in column M if available
                    if 'lithology_qualifier' in unit and unit['lithology_qualifier']:
                        safe_write_cell(sheet, f'{column_mapping["lithology_qualifier"]}{row_num}', unit['lithology_qualifier'])

                    # Write shade in column N if available in the units dataframe
                    if 'shade' in unit and unit['shade']:
                        safe_write_cell(sheet, f'{column_mapping["shade"]}{row_num}', unit['shade'])
                    
                    # Write hue in column O if available
                    if 'hue' in unit and unit['hue']:
                        safe_write_cell(sheet, f'{column_mapping["hue"]}{row_num}', unit['hue'])
                    
                    # Write colour in column P if available
                    if 'colour' in unit and unit['colour']:
                        safe_write_cell(sheet, f'{column_mapping["colour"]}{row_num}', unit['colour'])

                    # Write weathering in column Q if available
                    if 'weathering' in unit and unit['weathering']:
                        safe_write_cell(sheet, f'{column_mapping["weathering"]}{row_num}', unit['weathering'])

                    # Write estimated strength in column R if available
                    if 'estimated_strength' in unit and unit['estimated_strength']:
                        safe_write_cell(sheet, f'{column_mapping["estimated_strength"]}{row_num}', unit['estimated_strength'])

                    # Write record sequence in column S if available
                    if 'record_sequence' in unit and unit['record_sequence']:
                        safe_write_cell(sheet, f'{column_mapping["record_sequence"]}{row_num}', unit['record_sequence'])

                    # Write inter-relationship in column T if available
                    if 'inter_relationship' in unit and unit['inter_relationship']:
                        safe_write_cell(sheet, f'{column_mapping["inter_relationship"]}{row_num}', unit['inter_relationship'])

                    # Write percentage in column U if available
                    if 'percentage' in unit and unit['percentage'] > 0:
                        safe_write_cell(sheet, f'{column_mapping["percentage"]}{row_num}', unit['percentage'])

                    # Update progress every 100 units
                    if i % 100 == 0 and callback and i > 0:
                        callback(f"Writing unit {i+1} of {len(units)}...")

                # Debug: Log last 5 units being written
                if callback and len(units) > 5:
                     callback("DEBUG (save_to_template): Last 5 units to be written:")
                     for idx, unit_debug in units.tail(5).iterrows():
                         callback(f"  Index {idx}: {unit_debug.to_dict()}")

            # Update progress
            if callback:
                callback(f"Saving results to: {output_path}")
                
            # Save the workbook to the output path
            workbook.save(output_path)
            
            # Log success
            logger.info(f"Successfully saved results to {output_path} (Lithology sheet)")
            
            # Update progress
            if callback:
                callback(f"Results saved to Lithology sheet in: {output_path}")
                
            return True

        except PermissionError as pe:
            error_msg = f"PermissionError saving to template '{output_path}': {str(pe)}. Is the file open or write-protected?"
            logger.error(error_msg, exc_info=True) # Log traceback for permission errors
            if callback:
                callback(error_msg)
            return False
        except Exception as e:
            # Log the full traceback for unexpected errors
            error_msg = f"Unexpected error saving to template '{output_path}': {str(e)}"
            logger.error(error_msg, exc_info=True) # Log traceback for other errors
            if callback:
                callback(error_msg)
            return False
            # Save the workbook to the output path
